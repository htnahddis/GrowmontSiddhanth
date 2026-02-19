from django.db.models import Count
from rest_framework.decorators import api_view, permission_classes, parser_classes
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework import status
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework_simplejwt.views import TokenObtainPairView
from rest_framework_simplejwt.serializers import TokenObtainPairSerializer
from datetime import datetime

import openpyxl
from rest_framework.parsers import MultiPartParser, FormParser
from django.http import HttpResponse

from django.contrib.auth.models import User
from django.contrib.auth import authenticate
from django.contrib.auth.hashers import make_password
from django.core.mail import send_mail
from django.conf import settings
import random
import string

from .models import Employee, Sale, Interaction, Client, Reminder

from .serializers import (
    EmployeeSerializer,
    SaleSerializer,
    SaleCreateSerializer,
    InteractionCreateSerializer,
    InteractionSerializer,
    ClientSerializer,
    Interaction,
    ReminderSerializer
)


# ==================== CUSTOM JWT TOKEN ====================
class CustomTokenObtainPairSerializer(TokenObtainPairSerializer):
    def validate(self, attrs):
        try:
            data = super().validate(attrs)
        except Exception as e:
            # Log validation error for debugging
            print(f"Token validation error: {str(e)}")
            print(f"Attempted username: {attrs.get('username', 'N/A')}")
            raise
        
        # Add custom user data
        try:
            employee = self.user.employee
            # Safely get avatar URL
            avatar_url = None
            try:
                if employee.avatar and employee.avatar.name:
                    avatar_url = employee.avatar.url
            except Exception:
                avatar_url = None
            
            data['user'] = {
                'id': employee.id,
                'name': employee.name,
                'email': employee.email,
                'avatar': avatar_url,
                'role': employee.role,
            }
        except AttributeError:
            data['user'] = {
                'id': self.user.id,
                'name': self.user.username,
                'email': self.user.email,
                'avatar': None,
                'role': 'ADMIN',
            }
        
        return data


class CustomTokenObtainPairView(TokenObtainPairView):
    serializer_class = CustomTokenObtainPairSerializer


# ==================== HEALTH CHECK ====================
@api_view(["GET"])
def health_check(request):
    return Response({
        "status": "Backend running",
        "message": "Hello from Django"
    })


# ==================== AUTH PROTECTED TEST ====================
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def protected_view(request):
    return Response({
        "message": f"Hello {request.user.username}, you are authenticated"
    })


# ==================== LOGOUT ====================
class LogoutView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        try:
            refresh_token = request.data.get("refresh")
            if not refresh_token:
                return Response(
                    {"error": "Refresh token required"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            
            token = RefreshToken(refresh_token)
            token.blacklist()
            return Response(
                {"detail": "Successfully logged out"},
                status=status.HTTP_205_RESET_CONTENT,
            )
        except Exception as e:
            return Response(
                {"error": "Invalid token"},
                status=status.HTTP_400_BAD_REQUEST,
            )


# ==================== CURRENT USER ====================
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def current_user(request):
    try:
        serializer = EmployeeSerializer(request.user.employee)
        return Response(serializer.data)
    except AttributeError:
        return Response({'error': 'User is not an employee'}, status=400)


# ==================== CHANGE PASSWORD ====================
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def change_password(request):
    user = request.user
    old_password = request.data.get('old_password')
    new_password = request.data.get('new_password')

    if not user.check_password(old_password):
        return Response({'error': 'Wrong old password'}, status=400)

    user.set_password(new_password)
    user.save()
    return Response({'message': 'Password updated successfully'})


# ==================== FORGOT PASSWORD ====================
@api_view(['POST'])
def forgot_password(request):
    email = request.data.get('email')
    try:
        user = User.objects.get(email=email)
        # Generate temp password
        temp_password = ''.join(random.choices(string.ascii_letters + string.digits, k=8))
        user.set_password(temp_password)
        user.save()

        # Send Email
        print(f"========================================")
        print(f" PASSWORD RESET ")
        print(f" Email: {email}")
        print(f" Temp Password: {temp_password}")
        print(f"========================================")

        try:
            send_mail(
                'Password Reset Request',
                f'Hello,\n\nYour password has been reset.\n\nTemporary Password: {temp_password}\n\nPlease login and change it immediately.',
                'admin@growmont.com',
                [email],
                fail_silently=True,
            )
        except:
            pass
        
        return Response({'message': 'If an account exists, a new password has been sent.'})

    except User.DoesNotExist:
        # Return success even if user not found to prevent enumeration
        return Response({'message': 'If an account exists, a new password has been sent.'})


# ==================== EMPLOYEES ====================
@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def employees_list(request):
    # Check for Admin Role
    if hasattr(request.user, 'employee') and request.user.employee.role == 'EMPLOYEE':
        return Response({'error': 'Unauthorized'}, status=403)

    # CREATE EMPLOYEE
    if request.method == 'POST':
        data = request.data.copy()
        email = data.get('email')
        name = data.get('name')
        password = data.get('password')  # ADD THIS LINE

        # 1. Generate Password
        if not password:  # ADD THIS CHECK
            password = ''.join(random.choices(string.ascii_letters + string.digits, k=8))
            
        # 2. Create User
        try:
            user = User.objects.create_user(username=email, email=email, password=password)
        except Exception as e:
            return Response({'error': str(e)}, status=400)

        # 3. Create Employee
        serializer = EmployeeSerializer(data=data)
        if serializer.is_valid():
            employee = serializer.save(user=user)
            
            # # 4. Send Email
            # print(f"========================================")
            # print(f" NEW EMPLOYEE CREDENTIALS ")
            # print(f" Email: {email}")
            # print(f" Password: {password}")
            # print(f"========================================")

            # try:
            #     send_mail(
            #         'Your GrowMont Account Credentials',
            #         f'Hello {name},\n\nYour account has been created.\n\nUsername: {email}\nPassword: {password}\n\nPlease login and change your password.',
            #         'admin@growmont.com',
            #         [email],
            #         fail_silently=True,
            #     )
            # except:
            #     pass

            return Response(serializer.data, status=201)
        else:
            user.delete()  # Rollback user creation
            return Response(serializer.errors, status=400)

    # LIST EMPLOYEES
    employees = Employee.objects.annotate(
        clients_count=Count('clients', distinct=True),
        sales_count=Count('sales', distinct=True),
        interactions_count=Count('interactions', distinct=True),
    )
    serializer = EmployeeSerializer(employees, many=True)
    return Response(serializer.data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def employee_detail(request, id):
    try:
        employee = Employee.objects.annotate(
            clients_count=Count('clients', distinct=True),
            sales_count=Count('sales', distinct=True),
            interactions_count=Count('interactions', distinct=True),
        ).get(id=id)

        serializer = EmployeeSerializer(employee)
        return Response(serializer.data)
    except Employee.DoesNotExist:
        return Response({'error': 'Employee not found'}, status=404)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def employee_clients(request, id):
    clients = Client.objects.filter(employees__id=id).distinct()
    serializer = ClientSerializer(clients, many=True)
    return Response(serializer.data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def employee_sales(request, id):
    sales = Sale.objects.filter(sales_rep_id=id)
    serializer = SaleSerializer(sales, many=True)
    return Response(serializer.data)

@api_view(['PUT'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def update_employee(request, id):
    """Update employee - Admin only"""
    if hasattr(request.user, 'employee') and request.user.employee.role == 'EMPLOYEE':
        return Response({'error': 'Unauthorized'}, status=403)
    
    try:
        employee = Employee.objects.get(id=id)
    except Employee.DoesNotExist:
        return Response({'error': 'Employee not found'}, status=404)
    
    data = request.data.copy()
    
    # Don't update email if it's already taken by another employee
    if 'email' in data and data['email'] != employee.email:
        if Employee.objects.filter(email=data['email']).exclude(id=id).exists():
            return Response({'error': 'Email already exists'}, status=400)
    
    serializer = EmployeeSerializer(employee, data=data, partial=True)
    
    if serializer.is_valid():
        employee = serializer.save()
        return Response(EmployeeSerializer(employee).data, status=200)
    
    return Response(serializer.errors, status=400)


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_employee(request, id):
    """Delete employee - Admin only"""
    if hasattr(request.user, 'employee') and request.user.employee.role == 'EMPLOYEE':
        return Response({'error': 'Unauthorized'}, status=403)
    
    try:
        employee = Employee.objects.get(id=id)
        
        # Don't allow deleting yourself
        if hasattr(request.user, 'employee') and request.user.employee.id == id:
            return Response({'error': 'Cannot delete your own account'}, status=400)
        
        # Delete associated user
        if employee.user:
            employee.user.delete()
        
        employee.delete()
        return Response({'status': 'deleted'}, status=200)
        
    except Employee.DoesNotExist:
        return Response({'error': 'Employee not found'}, status=404)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def sales_list(request):
    user = request.user
    
    if hasattr(user, 'employee'):
        if user.employee.role == 'EMPLOYEE':
            sales = Sale.objects.filter(sales_rep=user.employee)
        else:
            sales = Sale.objects.all()
    else:
        sales = Sale.objects.all()
    
    serializer = SaleSerializer(sales, many=True)
    return Response(serializer.data)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_sale(request):
    data = request.data.copy()
    
    if hasattr(request.user, 'employee'):
        if request.user.employee.role == 'EMPLOYEE':
            data['sales_rep'] = request.user.employee.id
        else:
            if 'sales_rep' not in data or not data['sales_rep']:
                data['sales_rep'] = request.user.employee.id
    else:
        return Response({'error': 'User must be an employee'}, status=400)

    serializer = SaleCreateSerializer(data=data)
    
    if serializer.is_valid():
        sale = serializer.save()
        response_serializer = SaleSerializer(sale)
        return Response(response_serializer.data, status=201)
    
    return Response(serializer.errors, status=400)


@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def update_sale(request, pk):
    try:
        sale = Sale.objects.get(id=pk)
    except Sale.DoesNotExist:
        return Response({'error': 'Sale not found'}, status=404)
    
    user = request.user
    if hasattr(user, 'employee'):
        if user.employee.role == 'EMPLOYEE':
            if sale.sales_rep.id != user.employee.id:
                return Response({'error': 'Cannot edit other employees sales'}, status=403)
    
    data = request.data.copy()
    
    if hasattr(user, 'employee') and user.employee.role == 'EMPLOYEE':
        data['sales_rep'] = user.employee.id
    
    serializer = SaleCreateSerializer(sale, data=data, partial=True)
    
    if serializer.is_valid():
        sale = serializer.save()
        response_serializer = SaleSerializer(sale)
        return Response(response_serializer.data, status=200)
    
    return Response(serializer.errors, status=400)

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_sale(request, pk):
    try:
        Sale.objects.filter(id=pk).delete()
        return Response({'status': 'deleted'}, status=200)
    except Exception as e:
        return Response({'error': str(e)}, status=400)


# Add this to your existing views.py
# Replace the interactions section with this updated code

# ==================== INTERACTIONS (UPDATED) ====================
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def interactions_list(request):
    """
    List interactions based on user role
    - EMPLOYEE: Only their own interactions
    - ADMIN: All interactions
    """
    user = request.user
    
    if hasattr(user, 'employee'):
        if user.employee.role == 'EMPLOYEE':
            interactions = Interaction.objects.filter(employee=user.employee)
        else:
            interactions = Interaction.objects.all()
    else:
        interactions = Interaction.objects.all()
    
    serializer = InteractionSerializer(interactions, many=True)
    return Response(serializer.data)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_interaction(request):
    """
    Create new interaction
    - EMPLOYEE: Automatically assigned to themselves
    - ADMIN: Can assign to any employee
    """
    data = request.data.copy()
    
    # Determine which employee to assign
    if hasattr(request.user, 'employee'):
        if request.user.employee.role == 'EMPLOYEE':
            # Employee can only create for themselves
            data['employee'] = request.user.employee.id
        else:
            # Admin can assign to anyone
            if 'employee' not in data or not data['employee']:
                # Default to admin if not specified
                data['employee'] = request.user.employee.id
    else:
        return Response({'error': 'User must be an employee'}, status=400)

    serializer = InteractionCreateSerializer(data=data)
    
    if serializer.is_valid():
        interaction = serializer.save()
        
        # Return full interaction data
        response_serializer = InteractionSerializer(interaction)
        return Response(response_serializer.data, status=201)
    
    return Response(serializer.errors, status=400)


@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def update_interaction(request, pk):
    """
    Update interaction
    - EMPLOYEE: Can only update their own
    - ADMIN: Can update any
    """
    try:
        interaction = Interaction.objects.get(id=pk)
    except Interaction.DoesNotExist:
        return Response({'error': 'Interaction not found'}, status=404)
    
    # Check permissions
    user = request.user
    if hasattr(user, 'employee'):
        if user.employee.role == 'EMPLOYEE':
            if interaction.employee.id != user.employee.id:
                return Response({'error': 'Cannot edit other employees interactions'}, status=403)
    
    data = request.data.copy()
    
    # Employees cannot change the assigned employee
    if hasattr(user, 'employee') and user.employee.role == 'EMPLOYEE':
        data['employee'] = user.employee.id
    
    serializer = InteractionCreateSerializer(interaction, data=data, partial=True)
    
    if serializer.is_valid():
        interaction = serializer.save()
        response_serializer = InteractionSerializer(interaction)
        return Response(response_serializer.data, status=200)
    
    return Response(serializer.errors, status=400)


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_interaction(request, pk):
    """
    Delete interaction
    - EMPLOYEE: Can only delete their own
    - ADMIN: Can delete any
    """
    try:
        interaction = Interaction.objects.get(id=pk)
    except Interaction.DoesNotExist:
        return Response({'error': 'Interaction not found'}, status=404)
    
    # Check permissions
    user = request.user
    if hasattr(user, 'employee'):
        if user.employee.role == 'EMPLOYEE':
            if interaction.employee.id != user.employee.id:
                return Response({'error': 'Cannot delete other employees interactions'}, status=403)
    
    interaction.delete()
    return Response({'status': 'deleted'}, status=200)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def employees_for_dropdown(request):
    """
    Get list of employees for dropdown
    Used in interaction modal for admins
    """
    employees = Employee.objects.all().values('id', 'name', 'email', 'role')
    return Response(list(employees))


# ==================== EXPORT SALES ====================
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_sales_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Client', 'Date', 'Contact', 'Sales Rep', 'Product', 'Company', 'Scheme', 'Amount', 'Frequency', 'Remarks'])

    for s in Sale.objects.all():
        ws.append([
            s.client.name,
            s.date,
            s.client.contact_number,
            s.sales_rep.name,
            s.get_product_display(),
            s.company,
            s.scheme,
            s.amount,
            s.get_frequency_display(),
            s.remarks
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="sales.xlsx"'
    wb.save(response)
    return response


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def export_sales_excel_filtered(request):
    data = request.data.get('data', [])
    
    if not data:
        return Response({'error': 'No data provided'}, status=400)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Client', 'Date', 'Contact', 'Sales Rep', 'Product', 'Company', 'Scheme', 'Amount', 'Frequency', 'Remarks'])

    for row in data:
        ws.append([
            row.get('client', ''),
            row.get('date', ''),
            row.get('contactNo', ''),
            row.get('salesRep', ''),
            row.get('product', ''),
            row.get('company', ''),
            row.get('scheme', ''),
            row.get('amount', ''),
            row.get('frequency', ''),
            row.get('remark', '')
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="filtered_sales.xlsx"'
    wb.save(response)
    return response


# ==================== EXPORT INTERACTIONS ====================
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_interactions_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Client', 'Date', 'Contact', 'Notes', 'Next Follow Up'])

    for i in Interaction.objects.all():
        ws.append([
            i.client.name,
            i.date,
            i.client.contact_number,
            i.discussion_notes,
            i.next_follow_up
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="interactions.xlsx"'
    wb.save(response)
    return response


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def export_interactions_excel_filtered(request):
    data = request.data.get('data', [])
    
    if not data:
        return Response({'error': 'No data provided'}, status=400)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Client', 'Date', 'Contact', 'Summary', 'Follow Up Date'])

    for row in data:
        ws.append([
            row.get('client', ''),
            row.get('date', ''),
            row.get('contactNo', ''),
            row.get('summary', ''),
            row.get('followUpDate', '')
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="filtered_interactions.xlsx"'
    wb.save(response)
    return response


# ==================== IMPORT SALES ====================
@api_view(['POST'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser])
def import_sales_excel(request):
    try:
        wb = openpyxl.load_workbook(request.FILES['file'])
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            client_name, date, contact, rep_name, product, company, scheme, amount, frequency, remarks = row

            if not client_name:
                continue

            client_obj, _ = Client.objects.get_or_create(
                name=client_name,
                defaults={'contact_number': contact or ''}
            )
            
            try:
                rep_obj = Employee.objects.get(name=rep_name)
            except Employee.DoesNotExist:
                continue

            Sale.objects.create(
                client=client_obj,
                sales_rep=rep_obj,
                date=date,
                company=company,
                product=product,
                scheme=scheme or '',
                amount=amount,
                frequency=frequency,
                remarks=remarks or ''
            )

        return Response({'status': 'Sales imported successfully'})
    except Exception as e:
        return Response({'error': str(e)}, status=400)


# ==================== IMPORT INTERACTIONS ====================
@api_view(['POST'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser])
def import_interactions_excel(request):
    try:
        wb = openpyxl.load_workbook(request.FILES['file'])
        ws = wb.active

        # Use first employee as default
        default_employee = Employee.objects.first()

        if not default_employee:
            return Response(
                {'error': 'No employee found. Create employee first.'},
                status=400
            )

        for row in ws.iter_rows(min_row=2, values_only=True):
            client, date, contact, summary, follow_up = row

            if not client:
                continue

            client_obj, _ = Client.objects.get_or_create(
                name=client,
                defaults={'contact_number': contact or ''}
            )

            # Ensure date format
            if isinstance(date, datetime):
                date = date.date()
            if isinstance(follow_up, datetime):
                follow_up = follow_up.date()

            Interaction.objects.create(
                client=client_obj,
                employee=default_employee,
                date=date,
                discussion_notes=summary,
                next_follow_up=follow_up
            )

        return Response({'status': 'Interactions imported successfully'})
    except Exception as e:
        return Response({'error': str(e)}, status=400)
    

# In core/views.py

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def reminders_list(request):
    """List reminders for logged-in user"""
    if hasattr(request.user, 'employee'):
        reminders = Reminder.objects.filter(employee=request.user.employee)
    else:
        reminders = Reminder.objects.none()
    
    serializer = ReminderSerializer(reminders, many=True)
    return Response(serializer.data)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_reminder(request):
    """Create new reminder"""
    data = request.data.copy()
    
    if hasattr(request.user, 'employee'):
        data['employee'] = request.user.employee.id
    else:
        return Response({'error': 'User must be an employee'}, status=400)
    
    serializer = ReminderSerializer(data=data)
    
    if serializer.is_valid():
        reminder = serializer.save()
        return Response(ReminderSerializer(reminder).data, status=201)
    
    return Response(serializer.errors, status=400)

@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def update_reminder(request, pk):
    """Update reminder - users can only update their own"""
    try:
        reminder = Reminder.objects.get(id=pk)
    except Reminder.DoesNotExist:
        return Response({'error': 'Reminder not found'}, status=404)
    
    # Check if reminder belongs to logged-in user
    user = request.user
    if hasattr(user, 'employee'):
        if reminder.employee.id != user.employee.id:
            return Response(
                {'error': 'Cannot edit other users reminders'}, 
                status=403
            )
    
    data = request.data.copy()
    # Ensure employee stays the same
    data['employee'] = user.employee.id
    
    serializer = ReminderSerializer(reminder, data=data, partial=True)
    
    if serializer.is_valid():
        reminder = serializer.save()
        return Response(ReminderSerializer(reminder).data, status=200)
    
    return Response(serializer.errors, status=400)


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_reminder(request, pk):
    """Delete reminder"""
    try:
        reminder = Reminder.objects.get(id=pk)
        
        if hasattr(request.user, 'employee'):
            if reminder.employee.id != request.user.employee.id:
                return Response({'error': 'Cannot delete other users reminders'}, status=403)
        
        reminder.delete()
        return Response({'status': 'deleted'}, status=200)
    except Reminder.DoesNotExist:
        return Response({'error': 'Reminder not found'}, status=404)