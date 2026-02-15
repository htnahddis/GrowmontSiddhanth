from django.db.models import Count
from rest_framework.decorators import api_view, permission_classes, parser_classes
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework import status
from rest_framework_simplejwt.tokens import RefreshToken
from datetime import datetime

import openpyxl
from rest_framework.parsers import MultiPartParser
from django.http import HttpResponse

from django.contrib.auth.models import User
from django.contrib.auth import authenticate
from django.contrib.auth.hashers import make_password
from django.core.mail import send_mail
from django.conf import settings
import random
import string

from .models import Employee, Sale, Interaction, Client

from .serializers import (
    EmployeeSerializer,
    SaleSerializer,
    SaleCreateSerializer,
    InteractionSerializer,
    ClientSerializer,
)

# Health Check
@api_view(["GET"])
def health_check(request):
    return Response({
        "status": "Backend running",
        "message": "Hello from Django"
    })


# Auth Protected Test
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def protected_view(request):
    return Response({
        "message": f"Hello {request.user.username}, you are authenticated"
    })


# --------------------
# Logout
# --------------------
class LogoutView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        try:
            refresh_token = request.data["refresh"]
            token = RefreshToken(refresh_token)
            token.blacklist()
            return Response(
                {"detail": "Successfully logged out"},
                status=status.HTTP_205_RESET_CONTENT,
            )
        except Exception:
            return Response(
                {"error": "Invalid token"},
                status=status.HTTP_400_BAD_REQUEST,
            )


# --------------------
# Employees APIs  ✅ NEW
# --------------------
# @api_view(['GET'])
# def employees_list(request):
#     employees = Employee.objects.annotate(
#         clients_count=Count('clients', distinct=True),
#         sales_count=Count('sales', distinct=True),
#         interactions_count=Count('interactions', distinct=True),
#     )
#     serializer = EmployeeSerializer(employees, many=True)
#     return Response(serializer.data)

from rest_framework.parsers import MultiPartParser, FormParser

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def employees_list(request):

    # Check for Admin Role
    if hasattr(request.user, 'employee') and request.user.employee.role == 'EMPLOYEE':
        return Response({'error': 'Unauthorized'}, status=403)

    # 🔹 CREATE EMPLOYEE
    if request.method == 'POST':
        data = request.data.copy()
        email = data.get('email')
        name = data.get('name')

        # 1. Generate Password
        password = ''.join(random.choices(string.ascii_letters + string.digits, k=8))
        
        # 2. Create User
        try:
            user = User.objects.create_user(username=email, email=email, password=password)
        except Exception as e:
            return Response({'error': str(e)}, status=400)

        # 3. Create Employee
        # serializer = EmployeeSerializer(data=data)
        # We need to manually handle this or use serializer. 
        # Since we have a OneToOne, we can just save the employee with the user.
        
        # Using serializer is better for validation of other fields
        serializer = EmployeeSerializer(data=data)
        if serializer.is_valid():
            employee = serializer.save(user=user)
            
            # 4. Send Email (Mocking for now if no settings)
            print(f"========================================")
            print(f" NEW EMPLOYEE CREDENTIALS ")
            print(f" Email: {email}")
            print(f" Password: {password}")
            print(f"========================================")

            try:
                send_mail(
                    'Your GrowMont Account Credentials',
                    f'Hello {name},\n\nYour account has been created.\n\nUsername: {email}\nPassword: {password}\n\nPlease login and change your password.',
                    'admin@growmont.com',
                    [email],
                    fail_silently=True,
                )
            except:
                pass

            return Response(serializer.data, status=201)
        else:
            user.delete() # Rollback user creation
            return Response(serializer.errors, status=400)

    # 🔹 LIST EMPLOYEES
    employees = Employee.objects.annotate(
        clients_count=Count('clients', distinct=True),
        sales_count=Count('sales', distinct=True),
        interactions_count=Count('interactions', distinct=True),
    )
    serializer = EmployeeSerializer(employees, many=True)
    return Response(serializer.data)

@api_view(['GET'])
def employee_detail(request, id):
    employee = Employee.objects.annotate(
        clients_count=Count('clients', distinct=True),
        sales_count=Count('sales', distinct=True),
        interactions_count=Count('interactions', distinct=True),
    ).get(id=id)

    serializer = EmployeeSerializer(employee)
    return Response(serializer.data)


@api_view(['GET'])
def employee_clients(request, id):
    clients = Client.objects.filter(employees__id=id).distinct()
    serializer = ClientSerializer(clients, many=True)
    return Response(serializer.data)


@api_view(['GET'])
def employee_sales(request, id):
    sales = Sale.objects.filter(sales_rep_id=id)
    serializer = SaleSerializer(sales, many=True)
    return Response(serializer.data)


# --------------------
# Existing APIs
# --------------------
# @api_view(['GET'])
# def sales_list(request):
#     sales = Sale.objects.all()
#     serializer = SaleSerializer(sales, many=True)
#     return Response(serializer.data)
@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def sales_list(request):
    # GET: List all sales
    if request.method == 'GET':
        user = request.user
        if hasattr(user, 'employee') and user.employee.role == 'EMPLOYEE':
            sales = Sale.objects.filter(sales_rep=user.employee)
        else:
            sales = Sale.objects.all()
            
        serializer = SaleSerializer(sales, many=True)
        return Response(serializer.data)

    # POST: Create a sale
    if request.method == 'POST':
        data = request.data
    
        # CREATE CLIENT BY NAME
        client_name = data.pop('client_name', '')
        contact_no = data.pop('contactNo', '')
        
        if client_name:
            client, created = Client.objects.get_or_create(
                name=client_name,
                defaults={'contact_number': contact_no}
            )
            data['client'] = client.id
        else:
            return Response({'error': 'client_name required'}, status=400)
        
        # Default to first employee if not provided (fallback)
        if 'sales_rep' not in data:
             data['sales_rep'] = Employee.objects.first().id
             
        serializer = SaleCreateSerializer(data=data)
        if serializer.is_valid():
            sale = serializer.save()
            read_serializer = SaleSerializer(sale)
            return Response(read_serializer.data, status=201)
        return Response(serializer.errors, status=400)



@api_view(['GET'])
@permission_classes([IsAuthenticated])
def interactions_list(request):
    user = request.user
    if hasattr(user, 'employee') and user.employee.role == 'EMPLOYEE':
        interactions = Interaction.objects.filter(employee=user.employee)
    else:
        interactions = Interaction.objects.all()
    serializer = InteractionSerializer(interactions, many=True)
    return Response(serializer.data)


# export sales 
@api_view(['GET'])
def export_sales_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Client', 'Date', 'Contact', 'Sales Rep', 'Company', 'Amount', 'Remarks'])

    for s in Sale.objects.all():
        ws.append([
            s.client.name,
            s.date,
            s.client.contact_number,
            s.sales_rep.name,
            s.company,
            s.amount,
            s.remarks
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="sales.xlsx"'
    wb.save(response)
    return response


# Filtered Export Sales
@api_view(['POST'])
def export_sales_excel_filtered(request):
    data = request.data.get('data', [])

    print("EXPORT SALES HIT", request.data)
    
    if not data:
        return Response({'error': 'No data provided'}, status=400)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Client', 'Date', 'Contact', 'Sales Rep', 'Company', 'Amount', 'Remarks'])

    for row in data:
        ws.append([
            row.get('client', ''),
            row.get('date', ''),
            row.get('contactNo', ''),
            row.get('salesRep', ''),
            row.get('company', ''),
            row.get('amount', ''),
            row.get('remark', '')
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="filtered_sales.xlsx"'
    wb.save(response)
    return response


# Filtered Export Interactions
@api_view(['POST'])
def export_interactions_excel_filtered(request):
    data = request.data.get('data', [])
    
    if not data:
        return Response({'error': 'No data provided'}, status=400)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Client', 'Date', 'Contact', 'Summary', 'Follow Up Date'])

    for row in data:
        ws.append([
            row.get('client', ''),
            row.get('date', ''),
            row.get('contactNo', ''),
            row.get('summary', ''),
            row.get('followUpDate', '')
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="filtered_interactions.xlsx"'
    wb.save(response)
    return response


# export interactions 
@api_view(['GET'])
def export_interactions_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Client', 'Date', 'Contact', 'Notes', 'Next Follow Up'])

    for i in Interaction.objects.all():
        ws.append([
            i.client.name,
            i.date,
            i.client.contact_number,
            i.discussion_notes,
            i.next_follow_up
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="interactions.xlsx"'
    wb.save(response)
    return response



# import sales 
@api_view(['POST'])
@parser_classes([MultiPartParser])
def import_sales_excel(request):
    wb = openpyxl.load_workbook(request.FILES['file'])
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        client, date, contact, rep, company, amount, remarks = row

        client_obj, _ = Client.objects.get_or_create(
            name=client,
            contact_number=contact
        )
        rep_obj = Employee.objects.get(name=rep)

        Sale.objects.create(
            client=client_obj,
            sales_rep=rep_obj,
            date=date,
            company=company,
            amount=amount,
            remarks=remarks
        )

    return Response({'status': 'Sales imported'})


# import interactions
@api_view(['POST'])
@parser_classes([MultiPartParser])
def import_interactions_excel(request):
    wb = openpyxl.load_workbook(request.FILES['file'])
    ws = wb.active

    # TEMP: assign interactions to first employee (or manager)
    default_employee = Employee.objects.first()

    if not default_employee:
        return Response(
            {'error': 'No employee found. Create employee first.'},
            status=400
        )

    for row in ws.iter_rows(min_row=2, values_only=True):
        client, date, contact, summary, follow_up = row

        if not client:
            continue

        client_obj, _ = Client.objects.get_or_create(
            name=client,
            contact_number=contact
        )

        # Ensure date format
        if isinstance(date, datetime):
            date = date.date()
        if isinstance(follow_up, datetime):
            follow_up = follow_up.date()

        Interaction.objects.create(
            client=client_obj,
            employee=default_employee,      # ✅ REQUIRED
            date=date,
            discussion_notes=summary,
            next_follow_up=follow_up
        )

    return Response({'status': 'Interactions imported successfully'})


#create sales
@api_view(['POST'])
def create_sale(request):
    data = request.data

    client, _ = Client.objects.get_or_create(
        name=data['client'],
        defaults={'contact_number': data.get('contactNo', '')}  # Use get() for optional fields
    )

    try:
        if request.user.employee.role == 'EMPLOYEE':
             employee = request.user.employee
        else:
             employee = Employee.objects.get(id=data['employeeId'])
    except Employee.DoesNotExist:
        return Response({'error': 'Employee not found'}, status=400)

    sale = Sale.objects.create(
        client=client,
        sales_rep=employee,
        date=data['date'],
        company=data['company'],
        amount=data['amount'],
        product=data['product'], # Added
        frequency=data['frequency'], # Added
        remarks=data.get('remark', '')
    )

    return Response({'id': sale.id}, status=201)


#create interactions
@api_view(['POST'])
def create_interaction(request):
    data = request.data

    client, _ = Client.objects.get_or_create(
        name=data['client'],
        contact_number=data['contactNo']
    )

    employee = Employee.objects.get(id=data['employeeId'])

    Interaction.objects.create(
        client=client,
        employee=employee,
        date=data['date'],
        discussion_notes=data['summary'],
        next_follow_up=data['followUpDate']
    )

    return Response({'status': 'created'}, status=201)


# Update Sale
@api_view(['PUT'])
def update_sale(request, pk):
    try:
        sale = Sale.objects.get(id=pk)
    except Sale.DoesNotExist:
        return Response({'error': 'Sale not found'}, status=404)
    
    data = request.data
    
    # Update client if changed
    if 'client' in data and 'contactNo' in data:
        client, _ = Client.objects.get_or_create(
            name=data['client'],
            contact_number=data['contactNo']
        )
        sale.client = client
    
    # Update other fields
    if 'employeeId' in data:
        sale.sales_rep = Employee.objects.get(id=data['employeeId'])
    if 'date' in data:
        sale.date = data['date']
    if 'company' in data:
        sale.company = data['company']
    if 'amount' in data:
        sale.amount = data['amount']
    if 'remark' in data:
        sale.remarks = data['remark']
    
    sale.save()
    return Response({'status': 'updated'}, status=200)


# Update Interaction
@api_view(['PUT'])
def update_interaction(request, pk):
    try:
        interaction = Interaction.objects.get(id=pk)
    except Interaction.DoesNotExist:
        return Response({'error': 'Interaction not found'}, status=404)
    
    data = request.data
    
    # Update client if changed
    if 'client' in data and 'contactNo' in data:
        client, _ = Client.objects.get_or_create(
            name=data['client'],
            contact_number=data['contactNo']
        )
        interaction.client = client
    
    # Update other fields
    if 'employeeId' in data:
        interaction.employee = Employee.objects.get(id=data['employeeId'])
    if 'date' in data:
        interaction.date = data['date']
    if 'summary' in data:
        interaction.discussion_notes = data['summary']
    if 'followUpDate' in data:
        interaction.next_follow_up = data['followUpDate']
    
    interaction.save()
    return Response({'status': 'updated'}, status=200)


@api_view(['DELETE'])
def delete_sale(request, pk):
    Sale.objects.filter(id=pk).delete()
    return Response({'status': 'deleted'})

@api_view(['DELETE'])
def delete_interaction(request, pk):
    Interaction.objects.filter(id=pk).delete()
    return Response({'status': 'deleted'})

# core/views.py
@api_view(['POST'])
@parser_classes([MultiPartParser])
def create_employee(request):
    serializer = EmployeeSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=201)
    return Response(serializer.errors, status=400)


# Change Password
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def change_password(request):
    user = request.user
    old_password = request.data.get('old_password')
    new_password = request.data.get('new_password')

    if not user.check_password(old_password):
        return Response({'error': 'Wrong old password'}, status=400)

    user.set_password(new_password)
    user.save()
    return Response({'message': 'Password updated successfully'})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def current_user(request):
    try:
        serializer = EmployeeSerializer(request.user.employee)
        return Response(serializer.data)
    except AttributeError:
        return Response({'error': 'User is not an employee'}, status=400)


@api_view(['POST'])
def forgot_password(request):
    email = request.data.get('email')
    try:
        user = User.objects.get(email=email)
        # Generate temp password
        temp_password = ''.join(random.choices(string.ascii_letters + string.digits, k=8))
        user.set_password(temp_password)
        user.save()

        # Send Email
        print(f"========================================")
        print(f" PASSWORD RESET ")
        print(f" Email: {email}")
        print(f" Temp Password: {temp_password}")
        print(f"========================================")

        try:
            send_mail(
                'Password Reset Request',
                f'Hello,\n\nYour password has been reset.\n\nTemporary Password: {temp_password}\n\nPlease login and change it immediately.',
                'admin@growmont.com',
                [email],
                fail_silently=True,
            )
        except:
            pass
        
        return Response({'message': 'If an account exists, a new password has been sent.'})

    except User.DoesNotExist:
        # Return success even if user not found to prevent enumeration
        return Response({'message': 'If an account exists, a new password has been sent.'})

